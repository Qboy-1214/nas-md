import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from collections import defaultdict

# Load reward data
with open(r"C:\Users\Yong\Desktop\OH-WorkSpace\m42_rewards_v3.json", "r", encoding="utf-8") as f:
    data = json.load(f)

rewards = data["Sheet1"]

# Separate task rewards and chapter rewards
task_rewards = []
chapter_rewards = []
in_chapter = False

for row in rewards[1:]:
    if row[0] == "章节奖励":
        in_chapter = True
        continue
    if not in_chapter:
        if row[3] and str(row[3]).strip():
            task_rewards.append(row)
    else:
        if row[0] and str(row[0]).strip():
            chapter_rewards.append(row)

# Create workbook
wb = openpyxl.Workbook()

# ============================================================
# Sheet 1
# ============================================================
ws1 = wb.active
ws1.title = "TestCases"

headers = ["No", "Priority", "Module", "Type", "Precondition", "Operation", "Expected", "Result"]
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

p1_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
p2_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

cases = []
case_id = 1

for i, row in enumerate(task_rewards):
    chapter = row[1] if row[1] else ""
    seq = int(row[2]) if row[2] else 0
    task_desc = str(row[3]).strip() if row[3] else ""
    item1 = str(row[4]).strip() if row[4] else ""
    qty1 = row[5] if row[5] else 0
    item2 = str(row[6]).strip() if row[6] else ""
    qty2 = row[7] if row[7] else 0
    item3 = str(row[8]).strip() if row[8] else ""
    qty3 = row[9] if row[9] else 0
    blueprint = str(row[10]).strip() if row[10] else ""
    exp = row[11] if row[11] else 0
    rc = row[12] if row[12] else 0
    crystal = row[13] if row[13] else 0
    alloy = row[14] if row[14] else 0

    if not task_desc:
        continue

    module = f"Chapter {chapter}" if chapter else "Main"

    rewards_list = []
    if item1:
        rewards_list.append(f"{item1}x{int(qty1)}")
    if item2:
        rewards_list.append(f"{item2}x{int(qty2)}")
    if item3:
        rewards_list.append(f"{item3}x{int(qty3)}")
    if blueprint:
        rewards_list.append(blueprint)
    rewards_list.append(f"EXP{int(exp)}")
    if rc:
        rewards_list.append(f"RC{int(rc)}")
    if crystal:
        rewards_list.append(f"Crystal{int(crystal)}")
    if alloy:
        rewards_list.append(f"Alloy{int(alloy)}")

    expected = f"Got: {', '.join(rewards_list)}" if rewards_list else "No reward"

    cases.append(
        {
            "id": f"TC-{case_id:03d}",
            "priority": "P1",
            "module": module,
            "type": "Positive",
            "precondition": f"Complete {task_desc}",
            "operation": f"Complete {task_desc}",
            "expected": expected,
            "result": "",
        }
    )

    has_zero = (qty1 == 0 and item1) or (qty2 == 0 and item2) or (qty3 == 0 and item3)
    has_missing = not item1 and not item2 and not item3 and not blueprint

    if has_zero or has_missing:
        boundary_parts = []
        if has_missing:
            boundary_parts.append("no item/blueprint")
        if has_zero:
            boundary_parts.append("zero quantity")

        cases.append(
            {
                "id": f"TC-{case_id:03d}-B",
                "priority": "P2",
                "module": module,
                "type": "Boundary",
                "precondition": f"Complete {task_desc}",
                "operation": f'Complete task, check {", ".join(boundary_parts)}',
                "expected": "System handles empty/zero values correctly",
                "result": "",
            }
        )

    case_id += 1

for i, row in enumerate(chapter_rewards):
    chapter = str(row[0]).strip() if row[0] else ""
    gec = row[3] if row[3] else 0

    items = []
    for j in range(4, 14, 2):
        if j < len(row) and row[j] and str(row[j]).strip():
            item_name = str(row[j]).strip()
            item_qty = row[j + 1] if j + 1 < len(row) and row[j + 1] else 0
            items.append(f"{item_name}x{int(item_qty)}")

    expected = f"Got GEC{int(gec)}"
    if items:
        expected += f", {', '.join(items)}"

    cases.append(
        {
            "id": f"TC-C{i+1:03d}",
            "priority": "P1",
            "module": f"Chapter {chapter} Reward",
            "type": "Positive",
            "precondition": f"Complete all Chapter {chapter} tasks",
            "operation": f"Complete Chapter {chapter}",
            "expected": expected,
            "result": "",
        }
    )

for idx, case in enumerate(cases, 2):
    priority = case["priority"]
    row_fill = p1_fill if priority == "P1" else p2_fill

    values = [
        case["id"],
        priority,
        case["module"],
        case["type"],
        case["precondition"],
        case["operation"],
        case["expected"],
        case["result"],
    ]

    for col, val in enumerate(values, 1):
        cell = ws1.cell(row=idx, column=col, value=val)
        cell.fill = row_fill
        cell.border = thin_border
        cell.font = Font(name="Microsoft YaHei", size=10)
        if col in [1, 2, 3, 4, 8]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Data validation
dv = DataValidation(type="list", formula1='"P,F,?"', allow_blank=True)
dv.error = "Only P, F, or ? allowed"
ws1.add_data_validation(dv)
dv.add(f"H2:H{len(cases)+1}")

# Conditional formatting - simplified for Excel compatibility
# Rule 1: P = green
rule_p = CellIsRule(
    operator="equal",
    formula=['"P"'],
    fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
)
ws1.conditional_formatting.add(f"H2:H{len(cases)+1}", rule_p)
# Rule 2: F = red
rule_f = CellIsRule(
    operator="equal",
    formula=['"F"'],
    fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
)
ws1.conditional_formatting.add(f"H2:H{len(cases)+1}", rule_f)
# Rule 3: ? = yellow
rule_q = CellIsRule(
    operator="equal",
    formula=['"?"'],
    fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
)
ws1.conditional_formatting.add(f"H2:H{len(cases)+1}", rule_q)

col_widths = [18, 10, 22, 12, 35, 40, 55, 12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:H{len(cases)+1}"

# ============================================================
# Sheet 2: Coverage Matrix
# ============================================================
ws2 = wb.create_sheet("Coverage")

matrix_headers = ["Module", "Total", "Positive", "Negative", "Boundary", "Full Coverage"]
for col, h in enumerate(matrix_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

module_stats = defaultdict(lambda: {"total": 0, "Positive": 0, "Negative": 0, "Boundary": 0})
for case in cases:
    mod = case["module"]
    module_stats[mod]["total"] += 1
    module_stats[mod][case["type"]] += 1

row_idx = 2
for mod, stats in sorted(module_stats.items()):
    full = (
        "Yes" if stats["Positive"] > 0 and stats["Negative"] > 0 and stats["Boundary"] > 0 else "No"
    )
    values = [mod, stats["total"], stats["Positive"], stats["Negative"], stats["Boundary"], full]
    for col, val in enumerate(values, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.border = thin_border
        cell.font = Font(name="Microsoft YaHei", size=10)
        if col == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    row_idx += 1

matrix_widths = [25, 10, 10, 10, 10, 15]
for i, w in enumerate(matrix_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# ============================================================
# Sheet 3: Summary
# ============================================================
ws3 = wb.create_sheet("Summary")

ws3.merge_cells("A1:D1")
title_cell = ws3.cell(row=1, column=1, value="M42 H5 Main Quest Reward v3 - Test Case Summary")
title_cell.font = Font(name="Microsoft YaHei", size=14, bold=True)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

stat_headers = ["Item", "Count", "Percentage", "Note"]
for col, h in enumerate(stat_headers, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

total = len(cases)
p1_count = sum(1 for c in cases if c["priority"] == "P1")
p2_count = sum(1 for c in cases if c["priority"] == "P2")
pos_count = sum(1 for c in cases if c["type"] == "Positive")
neg_count = sum(1 for c in cases if c["type"] == "Negative")
bnd_count = sum(1 for c in cases if c["type"] == "Boundary")
mod_count = len(module_stats)

stats_data = [
    ("Total Cases", total, "100%", f"Cover {mod_count} modules"),
    ("P1 Important", p1_count, f"{p1_count/total*100:.1f}%", "Critical path"),
    ("P2 Normal", p2_count, f"{p2_count/total*100:.1f}%", "Can be added later"),
    ("Positive", pos_count, f"{pos_count/total*100:.1f}%", "Normal path"),
    ("Negative", neg_count, f"{neg_count/total*100:.1f}%", "Exception path"),
    ("Boundary", bnd_count, f"{bnd_count/total*100:.1f}%", "Edge cases"),
    ("Module Coverage", f"{mod_count}/{mod_count}", "100%", "All modules covered"),
]

for idx, (item, count, pct, note) in enumerate(stats_data, 4):
    values = [item, count, pct, note]
    for col, val in enumerate(values, 1):
        cell = ws3.cell(row=idx, column=col, value=val)
        cell.border = thin_border
        cell.font = Font(name="Microsoft YaHei", size=10)
        if col in [1, 4]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

stat_widths = [25, 12, 12, 40]
for i, w in enumerate(stat_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

output_path = r"C:\Users\Yong\Desktop\OH-WorkSpace\M42_H5_Main_Quest_Reward_v3_TestCases.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Total: {total}, Modules: {mod_count}")
print(f"P1: {p1_count}, P2: {p2_count}")
print(f"Positive: {pos_count}, Negative: {neg_count}, Boundary: {bnd_count}")
