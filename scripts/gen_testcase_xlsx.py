import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Load reward data
with open(r"C:\Users\Yong\Desktop\OH-WorkSpace\m42_rewards.json", "r", encoding="utf-8") as f:
    data = json.load(f)

rewards = data["Sheet1"]

# Header row
header = [
    "时间线",
    "章节",
    "序号",
    "",
    "道具1",
    "数量",
    "道具2",
    "数量",
    "部件蓝图",
    "经验奖励",
    "rc奖励",
    "水晶",
    "合金",
    "",
]

# Separate task rewards and chapter rewards
task_rewards = []
chapter_rewards = []
in_chapter = False

for row in rewards[1:]:  # Skip header
    if row[0] == "章节奖励":
        in_chapter = True
        continue
    if not in_chapter:
        # Only include rows that have a task description (column 3)
        if row[3] and str(row[3]).strip():
            task_rewards.append(row)
    else:
        if row[0] and str(row[0]).strip():
            chapter_rewards.append(row)

# Create workbook
wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: 测试用例总表
# ============================================================
ws1 = wb.active
ws1.title = "测试用例总表"

# Headers
headers = ["编号", "优先级", "功能模块", "用例类型", "前置条件", "操作步骤", "预期结果", "测试结果"]
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = header_border

# Priority fills
p0_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
p1_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
p2_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

# Data border
data_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# Generate test cases
cases = []
case_id = 1

# Task reward test cases
for i, row in enumerate(task_rewards):
    timeline = row[0] if row[0] else ""
    chapter = row[1] if row[1] else ""
    seq = int(row[2]) if row[2] else 0
    task_desc = str(row[3]).strip() if row[3] else ""
    item1 = str(row[4]).strip() if row[4] else ""
    qty1 = row[5] if row[5] else 0
    item2 = str(row[6]).strip() if row[6] else ""
    qty2 = row[7] if row[7] else 0
    blueprint = str(row[8]).strip() if row[8] else ""
    exp = row[9] if row[9] else 0
    rc = row[10] if row[10] else 0
    crystal = row[11] if row[11] else 0
    alloy = row[12] if row[12] else 0

    # Skip empty rows
    if not task_desc:
        continue

    # Determine module
    module = f"第{chapter}章" if chapter else "主线任务"

    # Positive case: reward values match
    rewards_list = []
    if item1:
        rewards_list.append(f"{item1}×{int(qty1)}")
    if item2:
        rewards_list.append(f"{item2}×{int(qty2)}")
    if blueprint:
        rewards_list.append(blueprint)
    rewards_list.append(f"经验{int(exp)}")
    if rc:
        rewards_list.append(f"RC{int(rc)}")
    if crystal:
        rewards_list.append(f"水晶{int(crystal)}")
    if alloy:
        rewards_list.append(f"合金{int(alloy)}")

    expected = f"获得: {', '.join(rewards_list)}" if rewards_list else "无奖励"

    # Positive case
    cases.append(
        {
            "id": f"TC-{case_id:03d}",
            "priority": "P1",
            "module": module,
            "type": "正向",
            "precondition": f"完成{task_desc}",
            "operation": f"完成{task_desc}任务",
            "expected": expected,
            "result": "",
        }
    )

    # Negative case: reward mismatch
    cases.append(
        {
            "id": f"TC-{case_id:03d}-N",
            "priority": "P2",
            "module": module,
            "type": "反向",
            "precondition": f"完成{task_desc}",
            "operation": f"完成{task_desc}任务后，修改本地奖励数值",
            "expected": "服务器返回正确奖励，本地修改无效",
            "result": "",
        }
    )

    # Boundary case: zero/missing fields
    if not item1 and not item2 and not blueprint:
        cases.append(
            {
                "id": f"TC-{case_id:03d}-B",
                "priority": "P2",
                "module": module,
                "type": "边界",
                "precondition": f"完成{task_desc}",
                "operation": f"完成任务，检查无道具奖励场景",
                "expected": "仅获得经验奖励，无道具/蓝图",
                "result": "",
            }
        )

    case_id += 1

# Chapter reward test cases
for i, row in enumerate(chapter_rewards):
    chapter = str(row[0]).strip() if row[0] else ""
    gec = row[3] if row[3] else 0

    # Parse items (columns 4-13, pairs of name+qty)
    items = []
    for j in range(4, 14, 2):
        if j < len(row) and row[j] and str(row[j]).strip():
            item_name = str(row[j]).strip()
            item_qty = row[j + 1] if j + 1 < len(row) and row[j + 1] else 0
            items.append(f"{item_name}×{int(item_qty)}")

    expected = f"获得gec{int(gec)}"
    if items:
        expected += f", {', '.join(items)}"

    cases.append(
        {
            "id": f"TC-C{i+1:03d}",
            "priority": "P1",
            "module": f"第{chapter}章完成奖励",
            "type": "正向",
            "precondition": f"完成第{chapter}章全部任务",
            "operation": f"完成第{chapter}章",
            "expected": expected,
            "result": "",
        }
    )

# Write cases to sheet
for idx, case in enumerate(cases, 2):
    priority = case["priority"]
    row_fill = p0_fill if priority == "P0" else (p1_fill if priority == "P1" else p2_fill)

    values = [
        case["id"],
        priority,
        case["module"],
        case["type"],
        case["precondition"],
        case["operation"],
        case["expected"],
        case["result"],
    ]

    for col, val in enumerate(values, 1):
        cell = ws1.cell(row=idx, column=col, value=val)
        cell.fill = row_fill
        cell.border = data_border
        cell.font = Font(name="微软雅黑", size=10)
        if col in [1, 2, 3, 4, 8]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Data validation for 测试结果 column (H)
dv = DataValidation(type="list", formula1='"P,F,？"', allow_blank=True)
dv.error = "只能填 P、F 或 ？"
dv.errorTitle = "输入无效"
ws1.add_data_validation(dv)
dv.add(f"H2:H{len(cases)+1}")

# Conditional formatting for 测试结果 column
from openpyxl.formatting.rule import CellIsRule

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

ws1.conditional_formatting.add(
    f"H2:H{len(cases)+1}",
    CellIsRule(
        operator="equal", formula=["P"], fill=green_fill, font=Font(color="006100", bold=True)
    ),
)
ws1.conditional_formatting.add(
    f"H2:H{len(cases)+1}",
    CellIsRule(
        operator="equal", formula=["F"], fill=red_fill, font=Font(color="9C0006", bold=True)
    ),
)
ws1.conditional_formatting.add(
    f"H2:H{len(cases)+1}",
    CellIsRule(
        operator="equal", formula=["？"], fill=yellow_fill, font=Font(color="9C6500", bold=True)
    ),
)

# Column widths
col_widths = [18, 8, 22, 10, 32, 42, 55, 12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws1.freeze_panes = "A2"

# Auto filter
ws1.auto_filter.ref = f"A1:H{len(cases)+1}"

# ============================================================
# Sheet 2: 覆盖度矩阵
# ============================================================
ws2 = wb.create_sheet("覆盖度矩阵")

matrix_headers = ["功能点", "覆盖用例数", "正向", "反向", "边界", "是否全覆盖"]
for col, h in enumerate(matrix_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border

# Count cases per module
from collections import defaultdict

module_stats = defaultdict(lambda: {"total": 0, "正向": 0, "反向": 0, "边界": 0})
for case in cases:
    mod = case["module"]
    module_stats[mod]["total"] += 1
    module_stats[mod][case["type"]] += 1

row_idx = 2
for mod, stats in sorted(module_stats.items()):
    full_coverage = "是" if stats["正向"] > 0 and stats["反向"] > 0 and stats["边界"] > 0 else "否"
    values = [mod, stats["total"], stats["正向"], stats["反向"], stats["边界"], full_coverage]
    for col, val in enumerate(values, 1):
        cell = ws2.cell(row=row_idx, column=col, value=val)
        cell.border = data_border
        cell.font = Font(name="微软雅黑", size=10)
        if col == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    row_idx += 1

# Column widths
matrix_widths = [24, 14, 8, 8, 8, 14]
for i, w in enumerate(matrix_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A2"

# ============================================================
# Sheet 3: 统计汇总
# ============================================================
ws3 = wb.create_sheet("统计汇总")

# Title
ws3.merge_cells("A1:D1")
title_cell = ws3.cell(row=1, column=1, value="M42 H5主线任务奖励 — 黑盒测试用例统计汇总")
title_cell.font = Font(name="微软雅黑", size=14, bold=True)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Headers
stat_headers = ["统计项", "数量", "占比", "说明"]
for col, h in enumerate(stat_headers, 1):
    cell = ws3.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = header_border

# Statistics
total = len(cases)
p0_count = sum(1 for c in cases if c["priority"] == "P0")
p1_count = sum(1 for c in cases if c["priority"] == "P1")
p2_count = sum(1 for c in cases if c["priority"] == "P2")
positive_count = sum(1 for c in cases if c["type"] == "正向")
negative_count = sum(1 for c in cases if c["type"] == "反向")
boundary_count = sum(1 for c in cases if c["type"] == "边界")
module_count = len(module_stats)

stats_data = [
    ("总用例数", total, "100%", f"覆盖 {module_count} 个功能点"),
    ("P0 核心功能", p0_count, f"{p0_count/total*100:.1f}%", "阻塞上线的关键路径"),
    ("P1 重要功能", p1_count, f"{p1_count/total*100:.1f}%", "影响体验的重要功能"),
    ("P2 一般功能", p2_count, f"{p2_count/total*100:.1f}%", "可后续补充"),
    ("正向用例", positive_count, f"{positive_count/total*100:.1f}%", "正常路径验证"),
    ("反向用例", negative_count, f"{negative_count/total*100:.1f}%", "异常路径验证"),
    ("边界用例", boundary_count, f"{boundary_count/total*100:.1f}%", "临界值和极端场景"),
    ("功能点覆盖率", f"{module_count}/{module_count}", "100%", "全部功能点均有对应用例"),
]

for idx, (item, count, pct, desc) in enumerate(stats_data, 4):
    values = [item, count, pct, desc]
    for col, val in enumerate(values, 1):
        cell = ws3.cell(row=idx, column=col, value=val)
        cell.border = data_border
        cell.font = Font(name="微软雅黑", size=10)
        if col in [1, 4]:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

# Column widths
stat_widths = [24, 12, 12, 40]
for i, w in enumerate(stat_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# Save
output_path = r"C:\Users\Yong\Desktop\OH-WorkSpace\M42_H5主线任务奖励_黑盒测试用例.xlsx"
wb.save(output_path)
print(f"Excel saved to: {output_path}")
print(f"Total cases: {total}")
print(f"Modules: {module_count}")
print(f"P0: {p0_count}, P1: {p1_count}, P2: {p2_count}")
print(f"Positive: {positive_count}, Negative: {negative_count}, Boundary: {boundary_count}")
